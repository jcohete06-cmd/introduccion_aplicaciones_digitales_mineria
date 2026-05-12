import streamlit as st
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
from scipy.stats import pearsonr
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import warnings
warnings.filterwarnings("ignore")

# ---------------------------- CONSTANTS (from original) ----------------------------
COLOR_HIST = "#2E86AB"
COLOR_CDF = "#A23B72"
COLOR_TORNADO_POS = "#2E86AB"
COLOR_TORNADO_NEG = "#E74C3C"
COLOR_SCATTER = "#2E86AB"
COLOR_DET = "#1B998B"
COLOR_DET2 = "#E74C3C"
COLOR_PIE = ["#2E86AB", "#A23B72", "#F18F01", "#C73E1D"]
COLOR_BAR = ["#2E86AB", "#A23B72", "#F18F01", "#1B998B"]

XL_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
XL_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
XL_SUB_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
XL_SUB_FONT = Font(bold=True, size=11)
XL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
XL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
XL_ORANGE = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
XL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
XL_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
XL_ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
XL_ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)

PARAM_INFO = {
    "x": {"label": "Block dimension X  (x)", "unit": "m", "hint": "Hard rock: 5-10 | Intermediate: 10-15 | Soft/bulk: 15-25", "default": 10.0, "std_frac": 0.0, "section": "BLOCK MODEL"},
    "y": {"label": "Block dimension Y  (y)", "unit": "m", "hint": "Hard rock: 5-10 | Intermediate: 10-15 | Soft/bulk: 15-25", "default": 10.0, "std_frac": 0.0, "section": "BLOCK MODEL"},
    "z": {"label": "Block dimension Z  (z)", "unit": "m", "hint": "Hard rock: 5-10 | Intermediate: 10-15 | Soft/bulk: 15-25", "default": 10.0, "std_frac": 0.0, "section": "BLOCK MODEL"},
    "rho_b": {"label": "Block density  (rho_b)", "unit": "t/m3", "hint": "Skarn Cu: 3.2-3.5 | Skarn Cu-Zn: 3.0-3.4 | Porphyry: 2.6-2.9 | Massive sulfide: 3.5-4.5", "default": 3.35, "std_frac": 0.03, "section": "BLOCK MODEL"},

    "g_Cu": {"label": "Cu grade  (g_Cu)", "unit": "%", "hint": "High grade: >1.5 | Moderate: 0.8-1.5 | Low: 0.3-0.8 | Very low: <0.3", "default": 1.14, "std_frac": 0.10, "section": "GRADES (Representative Block)"},
    "g_Zn": {"label": "Zn grade  (g_Zn)", "unit": "%", "hint": "Cu-ore: 0.1-0.3 | Cu-Zn ore: 1.5-3.0 | Zn dominant: >3.0", "default": 0.17, "std_frac": 0.15, "section": "GRADES (Representative Block)"},
    "g_Ag": {"label": "Ag grade  (g_Ag)", "unit": "g/t", "hint": "Low: <5 | Moderate: 5-15 | High: 15-30 | Very high: >30", "default": 8.7, "std_frac": 0.12, "section": "GRADES (Representative Block)"},
    "g_Mo": {"label": "Mo grade  (g_Mo)", "unit": "%", "hint": "Low: <0.01 | Moderate: 0.01-0.05 | High: 0.05-0.10", "default": 0.036, "std_frac": 0.15, "section": "GRADES (Representative Block)"},

    "C_min": {"label": "Mining cost  (C_min)", "unit": "USD/t", "hint": "Open pit large: 1.5-3.0 | Open pit medium: 3.0-5.0 | Underground: 5.0-15.0", "default": 2.10, "std_frac": 0.08, "section": "COSTS"},
    "C_proc": {"label": "Processing cost  (C_proc)", "unit": "USD/t", "hint": "Flotation simple: 5-8 | Flotation complex: 8-12 | Leaching: 3-6", "default": 7.50, "std_frac": 0.06, "section": "COSTS"},
    "C_vta": {"label": "Selling/refining cost  (C_vta)", "unit": "USD/t", "hint": "Domestic: 1.0-2.0 | Export: 2.0-4.0 | Remote: 4.0-8.0", "default": 1.80, "std_frac": 0.07, "section": "COSTS"},

    "P_Cu": {"label": "Cu price  (P_Cu)", "unit": "USD/t", "hint": "Bear: 6,000-7,500 | Base: 7,500-9,500 | Bull: 9,500-12,000", "default": 9000.0, "std_frac": 0.12, "section": "METAL PRICES"},
    "P_Zn": {"label": "Zn price  (P_Zn)", "unit": "USD/t", "hint": "Bear: 2,000-2,400 | Base: 2,400-3,200 | Bull: 3,200-4,000", "default": 2800.0, "std_frac": 0.12, "section": "METAL PRICES"},
    "P_Ag": {"label": "Ag price  (P_Ag)", "unit": "USD/t", "hint": "Bear: 20,000-25,000 | Base: 25,000-35,000 | Bull: 35,000-45,000 (~934 USD/oz)", "default": 30000.0, "std_frac": 0.15, "section": "METAL PRICES"},
    "P_Mo": {"label": "Mo price  (P_Mo)", "unit": "USD/t", "hint": "Bear: 25,000-35,000 | Base: 35,000-50,000 | Bull: 50,000-70,000", "default": 40000.0, "std_frac": 0.15, "section": "METAL PRICES"},

    "R_Cu": {"label": "Cu metallurgical recovery  (R_Cu)", "unit": "fraction", "hint": "Low refractory: 0.70-0.80 | Standard: 0.80-0.90 | Excellent: 0.90-0.95", "default": 0.87, "std_frac": 0.03, "section": "METALLURGICAL RECOVERY"},
    "R_Zn": {"label": "Zn metallurgical recovery  (R_Zn)", "unit": "fraction", "hint": "Low: 0.60-0.75 | Standard: 0.75-0.85 | Excellent: 0.85-0.92", "default": 0.80, "std_frac": 0.04, "section": "METALLURGICAL RECOVERY"},
    "R_Ag": {"label": "Ag metallurgical recovery  (R_Ag)", "unit": "fraction", "hint": "Byproduct low: 0.50-0.65 | Byproduct std: 0.65-0.80 | Primary: 0.80-0.95", "default": 0.75, "std_frac": 0.05, "section": "METALLURGICAL RECOVERY"},
    "R_Mo": {"label": "Mo metallurgical recovery  (R_Mo)", "unit": "fraction", "hint": "Separate circuit: 0.40-0.55 | Standard: 0.55-0.70 | Optimized: 0.70-0.80", "default": 0.50, "std_frac": 0.08, "section": "METALLURGICAL RECOVERY"},

    "D": {"label": "Dilution factor  (D)", "unit": "%", "hint": "Open pit selective: 3-5 | Open pit bulk: 5-10 | Underground SLS: 10-15 | Underground caving: 15-25", "default": 5.0, "std_frac": 0.10, "section": "DILUTION & MINING"},
    "g_w": {"label": "Waste grade Cu  (g_w)", "unit": "%", "hint": "Barren: 0.00-0.03 | Low grade: 0.03-0.10 | Mineralized waste: 0.10-0.30", "default": 0.05, "std_frac": 0.20, "section": "DILUTION & MINING"},
    "MR": {"label": "Mining recovery  (MR)", "unit": "fraction", "hint": "Open pit: 0.93-0.98 | UG room&piller: 0.80-0.90 | UG caving: 0.90-0.97", "default": 0.95, "std_frac": 0.02, "section": "DILUTION & MINING"},

    "T_Recurso_Medido": {"label": "Measured Resource tonnage  (T_Meas)", "unit": "Mt", "hint": "World-class: >500 | Large: 100-500 | Medium: 30-100 | Small: <30", "default": 400.0, "std_frac": 0.05, "section": "RESOURCES"},
    "g_Cu_Med": {"label": "Measured Cu grade", "unit": "%", "hint": "High: >1.2 | Moderate: 0.8-1.2 | Low: 0.4-0.8", "default": 1.10, "std_frac": 0.08, "section": "RESOURCES"},
    "g_Zn_Med": {"label": "Measured Zn grade", "unit": "%", "hint": "Cu-ore: 0.1-0.5 | Cu-Zn ore: 0.5-2.0 | Zn ore: >2.0", "default": 0.85, "std_frac": 0.10, "section": "RESOURCES"},
    "g_Ag_Med": {"label": "Measured Ag grade", "unit": "g/t", "hint": "Low: <5 | Moderate: 5-15 | High: >15", "default": 10.2, "std_frac": 0.10, "section": "RESOURCES"},
    "g_Mo_Med": {"label": "Measured Mo grade", "unit": "%", "hint": "Low: <0.01 | Moderate: 0.01-0.05 | High: >0.05", "default": 0.030, "std_frac": 0.12, "section": "RESOURCES"},

    "T_Recurso_Indicado": {"label": "Indicated Resource tonnage  (T_Ind)", "unit": "Mt", "hint": "World-class: >500 | Large: 100-500 | Medium: 30-100 | Small: <30", "default": 900.0, "std_frac": 0.08, "section": "RESOURCES"},
    "g_Cu_Ind": {"label": "Indicated Cu grade", "unit": "%", "hint": "High: >1.0 | Moderate: 0.6-1.0 | Low: 0.3-0.6", "default": 0.95, "std_frac": 0.10, "section": "RESOURCES"},
    "g_Zn_Ind": {"label": "Indicated Zn grade", "unit": "%", "hint": "Cu-ore: 0.1-0.5 | Cu-Zn ore: 0.5-2.0 | Zn ore: >2.0", "default": 0.75, "std_frac": 0.12, "section": "RESOURCES"},
    "g_Ag_Ind": {"label": "Indicated Ag grade", "unit": "g/t", "hint": "Low: <5 | Moderate: 5-12 | High: >12", "default": 9.5, "std_frac": 0.12, "section": "RESOURCES"},
    "g_Mo_Ind": {"label": "Indicated Mo grade", "unit": "%", "hint": "Low: <0.01 | Moderate: 0.01-0.04 | High: >0.04", "default": 0.025, "std_frac": 0.15, "section": "RESOURCES"},

    "T_Recurso_Inferido": {"label": "Inferred Resource tonnage  (T_Inf)", "unit": "Mt", "hint": "World-class: >500 | Large: 100-500 | Medium: 30-100 | Small: <30", "default": 634.0, "std_frac": 0.0, "section": "RESOURCES"},
    "g_Cu_Inf": {"label": "Inferred Cu grade", "unit": "%", "hint": "High: >0.8 | Moderate: 0.5-0.8 | Low: 0.2-0.5", "default": 0.75, "std_frac": 0.0, "section": "RESOURCES"},
    "g_Zn_Inf": {"label": "Inferred Zn grade", "unit": "%", "hint": "Cu-ore: 0.1-0.5 | Cu-Zn ore: 0.3-1.5 | Zn ore: >1.5", "default": 0.50, "std_frac": 0.0, "section": "RESOURCES"},
    "g_Ag_Inf": {"label": "Inferred Ag grade", "unit": "g/t", "hint": "Low: <5 | Moderate: 5-10 | High: >10", "default": 7.0, "std_frac": 0.0, "section": "RESOURCES"},
    "g_Mo_Inf": {"label": "Inferred Mo grade", "unit": "%", "hint": "Low: <0.01 | Moderate: 0.01-0.03 | High: >0.03", "default": 0.018, "std_frac": 0.0, "section": "RESOURCES"},

    "T_res_Cu_ore_P": {"label": "Reserve Proved Cu-ore tonnage", "unit": "Mt", "hint": "Main ore: >100 | Secondary: 20-100 | Minor: <20", "default": 100.0, "std_frac": 0.04, "section": "RESERVE TONNAGES"},
    "T_res_CuZn_ore_P": {"label": "Reserve Proved Cu-Zn ore tonnage", "unit": "Mt", "hint": "Main ore: >50 | Secondary: 10-50 | Minor: <10", "default": 42.0, "std_frac": 0.05, "section": "RESERVE TONNAGES"},
    "T_res_Cu_ore_Pr": {"label": "Reserve Probable Cu-ore tonnage", "unit": "Mt", "hint": "Main ore: >200 | Secondary: 50-200 | Minor: <50", "default": 454.0, "std_frac": 0.06, "section": "RESERVE TONNAGES"},
    "T_res_CuZn_ore_Pr": {"label": "Reserve Probable Cu-Zn ore tonnage", "unit": "Mt", "hint": "Main ore: >100 | Secondary: 30-100 | Minor: <30", "default": 149.0, "std_frac": 0.06, "section": "RESERVE TONNAGES"},

    "g_Cu_CuZn_P": {"label": "Proved Cu-Zn ore Cu grade", "unit": "%", "hint": "High: >1.2 | Moderate: 0.8-1.2 | Low: 0.4-0.8", "default": 0.99, "std_frac": 0.10, "section": "RESERVE GRADES: Proved Cu-Zn"},
    "g_Zn_CuZn_P": {"label": "Proved Cu-Zn ore Zn grade", "unit": "%", "hint": "Cu-Zn ore: 1.5-3.0 | Zn dominant: >3.0", "default": 2.30, "std_frac": 0.12, "section": "RESERVE GRADES: Proved Cu-Zn"},
    "g_Ag_CuZn_P": {"label": "Proved Cu-Zn ore Ag grade", "unit": "g/t", "hint": "Moderate: 10-20 | High: 20-40 | Very high: >40", "default": 19.5, "std_frac": 0.10, "section": "RESERVE GRADES: Proved Cu-Zn"},
    "g_Mo_CuZn_P": {"label": "Proved Cu-Zn ore Mo grade", "unit": "%", "hint": "Low: <0.01 | Moderate: 0.01-0.02", "default": 0.009, "std_frac": 0.15, "section": "RESERVE GRADES: Proved Cu-Zn"},

    "g_Cu_Cu_Pr": {"label": "Probable Cu-ore Cu grade", "unit": "%", "hint": "High: >1.2 | Moderate: 0.8-1.2 | Low: 0.4-0.8", "default": 1.05, "std_frac": 0.10, "section": "RESERVE GRADES: Probable Cu-ore"},
    "g_Zn_Cu_Pr": {"label": "Probable Cu-ore Zn grade", "unit": "%", "hint": "Cu-ore: 0.1-0.3 | Transitional: 0.3-1.0", "default": 0.17, "std_frac": 0.15, "section": "RESERVE GRADES: Probable Cu-ore"},
    "g_Ag_Cu_Pr": {"label": "Probable Cu-ore Ag grade", "unit": "g/t", "hint": "Low: <5 | Moderate: 5-12 | High: >12", "default": 9.7, "std_frac": 0.10, "section": "RESERVE GRADES: Probable Cu-ore"},
    "g_Mo_Cu_Pr": {"label": "Probable Cu-ore Mo grade", "unit": "%", "hint": "Low: <0.02 | Moderate: 0.02-0.04 | High: >0.04", "default": 0.031, "std_frac": 0.12, "section": "RESERVE GRADES: Probable Cu-ore"},

    "g_Cu_CuZn_Pr": {"label": "Probable Cu-Zn ore Cu grade", "unit": "%", "hint": "High: >1.0 | Moderate: 0.7-1.0 | Low: 0.4-0.7", "default": 1.05, "std_frac": 0.10, "section": "RESERVE GRADES: Probable Cu-Zn"},
    "g_Zn_CuZn_Pr": {"label": "Probable Cu-Zn ore Zn grade", "unit": "%", "hint": "Cu-Zn ore: 1.5-3.0 | Zn dominant: >3.0", "default": 2.07, "std_frac": 0.12, "section": "RESERVE GRADES: Probable Cu-Zn"},
    "g_Ag_CuZn_Pr": {"label": "Probable Cu-Zn ore Ag grade", "unit": "g/t", "hint": "Moderate: 10-20 | High: 20-40", "default": 17.9, "std_frac": 0.10, "section": "RESERVE GRADES: Probable Cu-Zn"},
    "g_Mo_CuZn_Pr": {"label": "Probable Cu-Zn ore Mo grade", "unit": "%", "hint": "Low: <0.01 | Moderate: 0.01-0.02", "default": 0.008, "std_frac": 0.15, "section": "RESERVE GRADES: Probable Cu-Zn"},

    "r_royalty": {"label": "Royalty rate  (r)", "unit": "fraction", "hint": "Peru: 0.01-0.03 | Chile: 0.04-0.09 | DRC: 0.06-0.10 | Australia: 0.05-0.08", "default": 0.03, "std_frac": 0.0, "section": "ROYALTY"},
    "N_sim": {"label": "Monte Carlo simulations  (N_sim)", "unit": "-", "hint": "Quick test: 5,000 | Standard: 10,000 | High precision: 50,000", "default": 10000, "std_frac": 0.0, "section": "MONTE CARLO"},
}

# ---------------------------- CORE CALCULATION FUNCTIONS ----------------------------
def sample_truncated_normal(mean, std_frac, n, lo=None, hi=None):
    if std_frac <= 0 or mean <= 0:
        return np.full(n, mean)
    std = mean * std_frac
    samples = np.random.normal(mean, std, n)
    if lo is None:
        lo = mean * 0.01
    if hi is None:
        hi = mean * 10
    samples = np.clip(samples, lo, hi)
    return samples

def calc_block_tonnage(x, y, z, rho_b):
    V_b = x * y * z
    T_b = V_b * rho_b
    return V_b, T_b

def calc_metal_content_base(T_total, g_pct):
    return T_total * g_pct / 100.0

def calc_metal_content_precious(T_total, g_gt):
    return T_total * g_gt / 1e6

def calc_cog(C_min, C_proc, C_vta, P_m, R_m, r):
    total_cost = C_min + C_proc + C_vta
    denom = P_m * R_m * (1 - r)
    if denom <= 0:
        return 0.0, total_cost
    cog_frac = total_cost / denom
    cog_pct = cog_frac * 100
    return cog_pct, total_cost

def calc_diluted_grade(g_is, D, g_w):
    return g_is * (1 - D / 100) + g_w * (D / 100)

def calc_rom_tonnage(T_res, D, MR):
    if (1 - D / 100) <= 0:
        return 0.0
    return T_res * MR / (1 - D / 100)

def calc_reserve_tonnage(T_res, MR, D):
    D_abs = D / (100 - D) if (100 - D) > 0 else 0
    return T_res * MR * (1 + D_abs)

def calc_reserve_grade(g_is, MR, g_w, T_res, T_R):
    if T_R <= 0:
        return 0.0
    return (g_is * MR + g_w * (T_R - T_res * MR) / T_R)

def calc_recoverable_metal_base(T_R, g_R, R_m):
    return T_R * g_R / 100 * R_m

def calc_recoverable_metal_precious(T_R, g_Ag, R_Ag):
    return T_R * g_Ag / 1e6 * R_Ag

def calc_cueq(g_Cu, g_Zn, g_Ag, g_Mo, P_Cu, P_Zn, P_Ag, P_Mo, R_Cu, R_Zn, R_Ag, R_Mo):
    V_Cu = g_Cu * 0.01 * P_Cu * R_Cu
    V_Zn = g_Zn * 0.01 * P_Zn * R_Zn
    V_Ag = g_Ag * 1e-6 * P_Ag * R_Ag
    V_Mo = g_Mo * 0.01 * P_Mo * R_Mo
    V_total = V_Cu + V_Zn + V_Ag + V_Mo
    denom = P_Cu * R_Cu * 0.01
    if denom <= 0:
        return 0.0, V_total, V_Cu, V_Zn, V_Ag, V_Mo
    CuEq = V_total / denom
    return CuEq, V_total, V_Cu, V_Zn, V_Ag, V_Mo

def calc_rf(T_Reserva, T_Recurso_MI):
    if T_Recurso_MI <= 0:
        return 0.0
    return T_Reserva / T_Recurso_MI * 100

def calc_margin(CuEq, P_Cu, R_Cu, C_min, C_proc, C_vta):
    I_t = (CuEq / 100) * P_Cu * R_Cu
    cost_total = C_min + C_proc + C_vta
    margin = I_t - cost_total
    return I_t, margin, cost_total

def calc_cueq_vec(mc):
    V_Cu = mc["g_Cu"] * 0.01 * mc["P_Cu"] * mc["R_Cu"]
    V_Zn = mc["g_Zn"] * 0.01 * mc["P_Zn"] * mc["R_Zn"]
    V_Ag = mc["g_Ag"] * 1e-6 * mc["P_Ag"] * mc["R_Ag"]
    V_Mo = mc["g_Mo"] * 0.01 * mc["P_Mo"] * mc["R_Mo"]
    V_total = V_Cu + V_Zn + V_Ag + V_Mo
    denom = mc["P_Cu"] * mc["R_Cu"] * 0.01
    denom = np.where(denom <= 0, 1e-10, denom)
    CuEq = V_total / denom
    return CuEq, V_total, V_Cu, V_Zn, V_Ag, V_Mo

def monte_carlo(p, N_sim, seed=42):
    np.random.seed(seed)
    mc = {}
    mc["rho_b"] = sample_truncated_normal(p["rho_b"], PARAM_INFO["rho_b"]["std_frac"], N_sim, 0.5, 6.0)
    mc["g_Cu"] = sample_truncated_normal(p["g_Cu"], PARAM_INFO["g_Cu"]["std_frac"], N_sim, 0.01, 10.0)
    mc["g_Zn"] = sample_truncated_normal(p["g_Zn"], PARAM_INFO["g_Zn"]["std_frac"], N_sim, 0.0, 15.0)
    mc["g_Ag"] = sample_truncated_normal(p["g_Ag"], PARAM_INFO["g_Ag"]["std_frac"], N_sim, 0.0, 200.0)
    mc["g_Mo"] = sample_truncated_normal(p["g_Mo"], PARAM_INFO["g_Mo"]["std_frac"], N_sim, 0.0, 1.0)
    mc["C_min"] = sample_truncated_normal(p["C_min"], PARAM_INFO["C_min"]["std_frac"], N_sim, 0.1, 30.0)
    mc["C_proc"] = sample_truncated_normal(p["C_proc"], PARAM_INFO["C_proc"]["std_frac"], N_sim, 0.5, 30.0)
    mc["C_vta"] = sample_truncated_normal(p["C_vta"], PARAM_INFO["C_vta"]["std_frac"], N_sim, 0.1, 15.0)
    mc["P_Cu"] = sample_truncated_normal(p["P_Cu"], PARAM_INFO["P_Cu"]["std_frac"], N_sim, 1000, 20000)
    mc["P_Zn"] = sample_truncated_normal(p["P_Zn"], PARAM_INFO["P_Zn"]["std_frac"], N_sim, 500, 8000)
    mc["P_Ag"] = sample_truncated_normal(p["P_Ag"], PARAM_INFO["P_Ag"]["std_frac"], N_sim, 5000, 80000)
    mc["P_Mo"] = sample_truncated_normal(p["P_Mo"], PARAM_INFO["P_Mo"]["std_frac"], N_sim, 5000, 100000)
    mc["R_Cu"] = sample_truncated_normal(p["R_Cu"], PARAM_INFO["R_Cu"]["std_frac"], N_sim, 0.3, 0.99)
    mc["R_Zn"] = sample_truncated_normal(p["R_Zn"], PARAM_INFO["R_Zn"]["std_frac"], N_sim, 0.2, 0.99)
    mc["R_Ag"] = sample_truncated_normal(p["R_Ag"], PARAM_INFO["R_Ag"]["std_frac"], N_sim, 0.1, 0.99)
    mc["R_Mo"] = sample_truncated_normal(p["R_Mo"], PARAM_INFO["R_Mo"]["std_frac"], N_sim, 0.1, 0.99)
    mc["D"] = sample_truncated_normal(p["D"], PARAM_INFO["D"]["std_frac"], N_sim, 0.0, 50.0)
    mc["g_w"] = sample_truncated_normal(p["g_w"], PARAM_INFO["g_w"]["std_frac"], N_sim, 0.0, 1.0)
    mc["MR"] = sample_truncated_normal(p["MR"], PARAM_INFO["MR"]["std_frac"], N_sim, 0.5, 1.0)
    mc["T_Recurso_Medido"] = sample_truncated_normal(p["T_Recurso_Medido"], PARAM_INFO["T_Recurso_Medido"]["std_frac"], N_sim, 1, 5000)
    mc["T_Recurso_Indicado"] = sample_truncated_normal(p["T_Recurso_Indicado"], PARAM_INFO["T_Recurso_Indicado"]["std_frac"], N_sim, 1, 5000)
    T_Reserva_total = p["T_res_Cu_ore_P"] + p["T_res_CuZn_ore_P"] + p["T_res_Cu_ore_Pr"] + p["T_res_CuZn_ore_Pr"]
    mc["T_Reserva"] = sample_truncated_normal(T_Reserva_total, 0.05, N_sim, 1, 5000)

    mc["COG"] = (mc["C_min"] + mc["C_proc"] + mc["C_vta"]) / (mc["P_Cu"] * mc["R_Cu"] * (1 - p["r_royalty"])) * 100
    mc["g_dil"] = mc["g_Cu"] * (1 - mc["D"] / 100) + mc["g_w"] * (mc["D"] / 100)
    mc["CuEq"], _, _, _, _, _ = calc_cueq_vec(mc)
    mc["I_t"] = (mc["CuEq"] / 100) * mc["P_Cu"] * mc["R_Cu"]
    mc["margin"] = mc["I_t"] - (mc["C_min"] + mc["C_proc"] + mc["C_vta"])
    mc["M_Cu_rec"] = mc["T_Reserva"] * 1e6 * mc["g_Cu"] / 100 * mc["R_Cu"]
    return mc

def mc_stats(arr):
    return {
        "mean": np.mean(arr),
        "std": np.std(arr),
        "P5": np.percentile(arr, 5),
        "P50": np.percentile(arr, 50),
        "P90": np.percentile(arr, 90),
        "P95": np.percentile(arr, 95),
    }

# ---------------------------- PLOTTING FUNCTIONS (return figures) ----------------------------
def plot_hist_cdf(data, title, xlabel):
    fig, ax1 = plt.subplots(figsize=(10, 6))
    n_bins = 50
    counts, bins, patches = ax1.hist(data, bins=n_bins, color=COLOR_HIST, alpha=0.7, edgecolor="white", linewidth=0.5, label="Histograma")
    ax1.set_xlabel(xlabel, fontsize=12)
    ax1.set_ylabel("Frecuencia", fontsize=12, color=COLOR_HIST)
    ax1.tick_params(axis="y", labelcolor=COLOR_HIST)
    ax2 = ax1.twinx()
    sorted_data = np.sort(data)
    cdf = np.arange(1, len(sorted_data) + 1) / len(sorted_data)
    ax2.plot(sorted_data, cdf, color=COLOR_CDF, linewidth=2.5, label="CDF acumulada")
    ax2.set_ylabel("Probabilidad acumulada", fontsize=12, color=COLOR_CDF)
    ax2.tick_params(axis="y", labelcolor=COLOR_CDF)
    for pct, lbl in [(5, "P5"), (50, "P50"), (95, "P95")]:
        val = np.percentile(data, pct)
        ax2.axvline(val, color="gray", linestyle="--", linewidth=0.8, alpha=0.7)
        ax2.annotate(f"{lbl}={val:.4f}", xy=(val, pct / 100), fontsize=8, color="gray",
                     xytext=(5, 5), textcoords="offset points")
    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2, loc="best", fontsize=10)
    ax1.set_title(title, fontsize=13, fontweight="bold")
    plt.tight_layout()
    return fig

def plot_tornado(mc, p):
    output_var = mc["margin"]
    input_keys = ["g_Cu", "P_Cu", "R_Cu", "C_proc", "C_min", "C_vta", "g_Zn", "P_Zn", "R_Zn",
                  "g_Ag", "P_Ag", "R_Ag", "g_Mo", "P_Mo", "R_Mo", "D", "MR"]
    correlations = {}
    for k in input_keys:
        if k in mc and len(mc[k]) == len(output_var):
            r_val, _ = pearsonr(mc[k], output_var)
            correlations[k] = r_val
    sorted_corr = sorted(correlations.items(), key=lambda x: abs(x[1]), reverse=True)
    labels = [item[0] for item in sorted_corr]
    values = [item[1] for item in sorted_corr]
    fig, ax = plt.subplots(figsize=(10, max(6, len(labels) * 0.4)))
    colors = [COLOR_TORNADO_POS if v >= 0 else COLOR_TORNADO_NEG for v in values]
    bars = ax.barh(range(len(labels)), values, color=colors, edgecolor="white", linewidth=0.5)
    ax.set_yticks(range(len(labels)))
    ax.set_yticklabels(labels, fontsize=10)
    ax.set_xlabel("Correlación de Pearson con Margen Operativo", fontsize=11)
    ax.set_title("Diagrama de Tornado: Sensibilidad del Margen Operativo", fontsize=13, fontweight="bold")
    ax.axvline(0, color="black", linewidth=0.8)
    for bar_item, val in zip(bars, values):
        x_pos = val + 0.01 if val >= 0 else val - 0.01
        ax.text(x_pos, bar_item.get_y() + bar_item.get_height() / 2, f"{val:.3f}",
                va="center", ha="left" if val >= 0 else "right", fontsize=8)
    ax.invert_yaxis()
    plt.tight_layout()
    return fig

def plot_scatter(mc, x_key, y_key, xlabel, ylabel, title):
    fig, ax = plt.subplots(figsize=(9, 7))
    ax.scatter(mc[x_key], mc[y_key], alpha=0.15, s=5, color=COLOR_SCATTER, edgecolors="none")
    z = np.polyfit(mc[x_key], mc[y_key], 1)
    poly = np.poly1d(z)
    x_line = np.linspace(mc[x_key].min(), mc[x_key].max(), 100)
    ax.plot(x_line, poly(x_line), color=COLOR_DET2, linewidth=2, label=f"Tendencia lineal")
    ax.set_xlabel(xlabel, fontsize=11)
    ax.set_ylabel(ylabel, fontsize=11)
    ax.set_title(title, fontsize=13, fontweight="bold")
    ax.legend(loc="best", fontsize=10)
    plt.tight_layout()
    return fig

def plot_det_sensitivity_price(p):
    prices = np.linspace(5000, 14000, 100)
    cogs = []
    margins = []
    for pc in prices:
        cog, _ = calc_cog(p["C_min"], p["C_proc"], p["C_vta"], pc, p["R_Cu"], p["r_royalty"])
        cueq, _, _, _, _, _ = calc_cueq(p["g_Cu"], p["g_Zn"], p["g_Ag"], p["g_Mo"],
                                         pc, p["P_Zn"], p["P_Ag"], p["P_Mo"],
                                         p["R_Cu"], p["R_Zn"], p["R_Ag"], p["R_Mo"])
        _, margin, _ = calc_margin(cueq, pc, p["R_Cu"], p["C_min"], p["C_proc"], p["C_vta"])
        cogs.append(cog)
        margins.append(margin)
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 6))
    ax1.plot(prices, cogs, color=COLOR_DET, linewidth=2.5)
    ax1.axhline(p["g_Cu"], color=COLOR_DET2, linestyle="--", linewidth=1.5, label=f"Ley Cu in situ ({p['g_Cu']:.2f}%)")
    ax1.set_xlabel("Precio Cu (USD/t)", fontsize=11)
    ax1.set_ylabel("Ley de Corte COG (% Cu)", fontsize=11)
    ax1.set_title("Precio Cu vs COG", fontsize=12, fontweight="bold")
    ax1.legend(loc="best", fontsize=10)
    ax1.grid(True, alpha=0.3)
    ax2.plot(prices, margins, color=COLOR_DET2, linewidth=2.5)
    ax2.axhline(0, color="black", linestyle="-", linewidth=0.8)
    ax2.set_xlabel("Precio Cu (USD/t)", fontsize=11)
    ax2.set_ylabel("Margen Operativo (USD/t)", fontsize=11)
    ax2.set_title("Precio Cu vs Margen", fontsize=12, fontweight="bold")
    ax2.grid(True, alpha=0.3)
    fig.suptitle("Sensibilidad Determinística al Precio del Cobre", fontsize=14, fontweight="bold", y=1.02)
    plt.tight_layout()
    return fig

def plot_det_sensitivity_grade(p):
    grades = np.linspace(0.3, 2.5, 100)
    cueqs = []
    margins = []
    for gc in grades:
        cueq, _, _, _, _, _ = calc_cueq(gc, p["g_Zn"], p["g_Ag"], p["g_Mo"],
                                          p["P_Cu"], p["P_Zn"], p["P_Ag"], p["P_Mo"],
                                          p["R_Cu"], p["R_Zn"], p["R_Ag"], p["R_Mo"])
        _, margin, _ = calc_margin(cueq, p["P_Cu"], p["R_Cu"], p["C_min"], p["C_proc"], p["C_vta"])
        cueqs.append(cueq)
        margins.append(margin)
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 6))
    ax1.plot(grades, cueqs, color=COLOR_DET, linewidth=2.5)
    ax1.set_xlabel("Ley Cu (%)", fontsize=11)
    ax1.set_ylabel("CuEq (%)", fontsize=11)
    ax1.set_title("Ley Cu vs CuEq", fontsize=12, fontweight="bold")
    ax1.grid(True, alpha=0.3)
    ax2.plot(grades, margins, color=COLOR_DET2, linewidth=2.5)
    ax2.axhline(0, color="black", linestyle="-", linewidth=0.8)
    ax2.set_xlabel("Ley Cu (%)", fontsize=11)
    ax2.set_ylabel("Margen Operativo (USD/t)", fontsize=11)
    ax2.set_title("Ley Cu vs Margen", fontsize=12, fontweight="bold")
    ax2.grid(True, alpha=0.3)
    fig.suptitle("Sensibilidad Determinística a la Ley de Cobre", fontsize=14, fontweight="bold", y=1.02)
    plt.tight_layout()
    return fig

def plot_resource_pie(p):
    labels = ["Medido", "Indicado", "Inferido"]
    sizes = [p["T_Recurso_Medido"], p["T_Recurso_Indicado"], p["T_Recurso_Inferido"]]
    total = sum(sizes)
    if total <= 0:
        fig, ax = plt.subplots()
        ax.text(0.5, 0.5, "No data", ha="center")
        return fig
    fig, ax = plt.subplots(figsize=(8, 8))
    wedges, texts, autotexts = ax.pie(sizes, labels=labels, autopct=lambda pct: f"{pct:.1f}%\n({pct*total/100:.0f} Mt)",
                                       colors=COLOR_PIE[:3], startangle=90, textprops={"fontsize": 11})
    for autotext in autotexts:
        autotext.set_fontsize(9)
    ax.set_title(f"Clasificación de Recursos (Total: {total:.0f} Mt)", fontsize=13, fontweight="bold")
    plt.tight_layout()
    return fig

def plot_metal_content_bar(reserve_data):
    categories = [d["name"] for d in reserve_data]
    M_Cu = [d["M_Cu"] for d in reserve_data]
    M_Zn = [d["M_Zn"] for d in reserve_data]
    M_Ag = [d["M_Ag"] for d in reserve_data]
    M_Mo = [d["M_Mo"] for d in reserve_data]
    fig, axes = plt.subplots(2, 2, figsize=(14, 10))
    for ax, vals, metal, unit in [
        (axes[0, 0], M_Cu, "Cu", "t"),
        (axes[0, 1], M_Zn, "Zn", "t"),
        (axes[1, 0], M_Ag, "Ag", "t"),
        (axes[1, 1], M_Mo, "Mo", "t"),
    ]:
        bars = ax.bar(range(len(categories)), vals, color=COLOR_BAR, edgecolor="white")
        ax.set_xticks(range(len(categories)))
        ax.set_xticklabels(categories, rotation=25, ha="right", fontsize=9)
        ax.set_ylabel(f"Contenido {metal} ({unit})", fontsize=10)
        ax.set_title(f"{metal} contenido in situ", fontsize=11, fontweight="bold")
        for bar_item, v in zip(bars, vals):
            ax.text(bar_item.get_x() + bar_item.get_width() / 2, bar_item.get_height(),
                    f"{v:,.0f}", ha="center", va="bottom", fontsize=8)
    fig.suptitle("Contenido Metálico por Categoría de Reserva", fontsize=14, fontweight="bold", y=1.02)
    plt.tight_layout()
    return fig

def plot_resource_reserve_comparison(p):
    categories = ["Medido", "Indicado", "Med+Ind", "Inferido"]
    resources = [p["T_Recurso_Medido"], p["T_Recurso_Indicado"],
                 p["T_Recurso_Medido"] + p["T_Recurso_Indicado"], p["T_Recurso_Inferido"]]
    reserves = [
        p["T_res_Cu_ore_P"] + p["T_res_CuZn_ore_P"],
        p["T_res_Cu_ore_Pr"] + p["T_res_CuZn_ore_Pr"],
        p["T_res_Cu_ore_P"] + p["T_res_CuZn_ore_P"] + p["T_res_Cu_ore_Pr"] + p["T_res_CuZn_ore_Pr"],
        0,
    ]
    x = np.arange(len(categories))
    width = 0.35
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.bar(x - width / 2, resources, width, label="Recurso (Mt)", color=COLOR_HIST, edgecolor="white")
    ax.bar(x + width / 2, reserves, width, label="Reserva (Mt)", color=COLOR_DET, edgecolor="white")
    ax.set_xticks(x)
    ax.set_xticklabels(categories, fontsize=11)
    ax.set_ylabel("Tonelaje (Mt)", fontsize=11)
    ax.set_title("Comparación Recurso vs Reserva por Categoría", fontsize=13, fontweight="bold")
    ax.legend(loc="best", fontsize=10)
    plt.tight_layout()
    return fig

def plot_cueq_by_category(p):
    cats = ["Medido", "Indicado", "Inferido"]
    cueqs = []
    for cat, suf in [("Medido", "Med"), ("Indicado", "Ind"), ("Inferido", "Inf")]:
        cueq, _, _, _, _, _ = calc_cueq(
            p[f"g_Cu_{suf}"], p[f"g_Zn_{suf}"], p[f"g_Ag_{suf}"], p[f"g_Mo_{suf}"],
            p["P_Cu"], p["P_Zn"], p["P_Ag"], p["P_Mo"],
            p["R_Cu"], p["R_Zn"], p["R_Ag"], p["R_Mo"],
        )
        cueqs.append(cueq)
    res_cats = ["Prob. Cu-ore", "Prob. Cu-Zn", "Prob. Cu-ore", "Prob. Cu-Zn"]
    res_cueqs = []
    for gcu, gzn, gag, gmo in [
        (p["g_Cu"], p["g_Zn"], p["g_Ag"], p["g_Mo"]),
        (p["g_Cu_CuZn_P"], p["g_Zn_CuZn_P"], p["g_Ag_CuZn_P"], p["g_Mo_CuZn_P"]),
        (p["g_Cu_Cu_Pr"], p["g_Zn_Cu_Pr"], p["g_Ag_Cu_Pr"], p["g_Mo_Cu_Pr"]),
        (p["g_Cu_CuZn_Pr"], p["g_Zn_CuZn_Pr"], p["g_Ag_CuZn_Pr"], p["g_Mo_CuZn_Pr"]),
    ]:
        cueq, _, _, _, _, _ = calc_cueq(gcu, gzn, gag, gmo,
                                         p["P_Cu"], p["P_Zn"], p["P_Ag"], p["P_Mo"],
                                         p["R_Cu"], p["R_Zn"], p["R_Ag"], p["R_Mo"])
        res_cueqs.append(cueq)
    all_cats = cats + res_cats
    all_cueqs = cueqs + res_cueqs
    fig, ax = plt.subplots(figsize=(11, 6))
    colors_all = COLOR_PIE[:3] + COLOR_BAR
    bars = ax.bar(range(len(all_cats)), all_cueqs, color=colors_all, edgecolor="white")
    ax.set_xticks(range(len(all_cats)))
    ax.set_xticklabels(all_cats, rotation=30, ha="right", fontsize=10)
    ax.set_ylabel("CuEq (%)", fontsize=11)
    ax.set_title("Equivalente de Cobre por Categoría", fontsize=13, fontweight="bold")
    for bar_item, v in zip(bars, all_cueqs):
        ax.text(bar_item.get_x() + bar_item.get_width() / 2, bar_item.get_height(),
                f"{v:.3f}", ha="center", va="bottom", fontsize=9)
    plt.tight_layout()
    return fig

# ---------------------------- INTERPRETATION FUNCTIONS ----------------------------
def risk_fill(level):
    return {"green": "🟢", "yellow": "🟡", "orange": "🟠", "red": "🔴"}.get(level, "")

def interpret_cog(cog):
    if cog < 0.10:
        level = "green"
        txt = (f"La ley de corte de equilibrio es {cog:.4f}% Cu ({cog*10000:.1f} ppm). "
               f"Este COG muy bajo indica que el depósito es altamente rentable bajo las condiciones actuales. "
               f"Comparado con el COG publicado de Antamina (~0.20% CuEq), el margen de seguridad es amplio. "
               f"Un COG < 0.10% es típico de depósitos de cobre de gran tonelaje y baja ley en operaciones a cielo abierto de gran escala. "
               f"Recomendación: Se puede considerar la extracción de bloques de menor ley, ampliando potencialmente la vida útil de la mina.")
    elif cog < 0.20:
        level = "green"
        txt = (f"La ley de corte de equilibrio es {cog:.4f}% Cu ({cog*10000:.1f} ppm). "
               f"Este COG es consistente con operaciones de tajo abierto de cobre a gran escala. "
               f"Es comparable al COG de diseño de Antamina (~0.20% CuEq). "
               f"El depósito muestra viabilidad económica razonable bajo las condiciones de costo y precio ingresadas. "
               f"Recomendación: Mantener el plan de minado actual; monitorear costos operativos para asegurar que el COG no se eleve por encima de las leyes marginales del pit.")
    elif cog < 0.40:
        level = "yellow"
        txt = (f"La ley de corte de equilibrio es {cog:.4f}% Cu ({cog*10000:.1f} ppm). "
               f"Este COG moderado indica que el depósito requiere leyes relativamente altas para ser económico. "
               f"Es típico de operaciones subterráneas o depósitos con costos de procesamiento elevados. "
               f"Parte del recurso de baja ley podría quedar fuera del pit económico. "
               f"Recomendación: Optimizar la secuencia de extracción para priorizar bloques de mayor ley. "
               f"Evaluar la reducción de costos de procesamiento o el aumento de recuperación metalúrgica.")
    elif cog < 0.80:
        level = "orange"
        txt = (f"La ley de corte de equilibrio es {cog:.4f}% Cu ({cog*10000:.1f} ppm). "
               f"Este COG alto indica una operación marginalmente rentable. "
               f"Solo los bloques con leyes significativamente superiores al promedio serán económicos. "
               f"Es típico de depósitos de cobre de baja ley con costos elevados o recuperaciones bajas. "
               f"Recomendación: Reevaluar la viabilidad del proyecto. Considerar alternativas de procesamiento de menor costo (lixiviación SX-EW). "
               f"Renegociar costos logísticos y de energía.")
    else:
        level = "red"
        txt = (f"La ley de corte de equilibrio es {cog:.4f}% Cu ({cog*10000:.1f} ppm). "
               f"Este COG excesivamente alto indica que la operación NO es económicamente viable bajo las condiciones actuales. "
               f"La gran mayoría del recurso no alcanzaría la ley mínima requerida. "
               f"Recomendación: El proyecto requiere cambios fundamentales: reducción drástica de costos, "
               f"aumento significativo del precio del metal, o mejora sustancial de la recuperación. "
               f"Se recomienda suspender la inversión hasta que las condiciones mejoren.")
    return txt, level

def interpret_cueq(cueq):
    if cueq > 1.5:
        level = "green"
        txt = (f"El equivalente de cobre es {cueq:.3f}% CuEq. "
               f"Esta ley CuEq es alta para un depósito polimetálico tipo skarn. "
               f"Indica que el valor económico combinado de todos los metales (Cu, Zn, Ag, Mo) es muy significativo. "
               f"La contribución de los subproductos (Zn, Ag, Mo) al valor total es notable, lo cual diversifica el riesgo de precio. "
               f"Recomendación: Continuar con el plan de explotación polimetálico. Evaluar la expansión del circuito de Mo si el precio lo justifica.")
    elif cueq > 0.8:
        level = "green"
        txt = (f"El equivalente de cobre es {cueq:.3f}% CuEq. "
               f"Esta ley CuEq es típica de depósitos de cobre de gran escala operando a cielo abierto. "
               f"El aporte de los metales accesorios (Zn, Ag, Mo) incrementa el valor por tonelada respecto al cobre solo. "
               f"Para Antamina, un CuEq de ~1.28% es consistente con la producción reportada de ~560,000 t CuEq/ano. "
               f"Recomendación: Mantener el enfoque polimetálico. Asegurar que los circuitos de flotación de Zn y Mo están optimizados.")
    elif cueq > 0.4:
        level = "yellow"
        txt = (f"El equivalente de cobre es {cueq:.3f}% CuEq. "
               f"Esta ley CuEq es moderada-baja. El depósito requiere volúmenes de procesamiento elevados para ser rentable. "
               f"La contribución de subproductos es probablemente baja o los precios relativos son desfavorables. "
               f"Recomendación: Evaluar la viabilidad de aumentar la capacidad de procesamiento. "
               f"Optimizar la recuperación de los metales accesorios para maximizar el CuEq.")
    else:
        level = "orange"
        txt = (f"El equivalente de cobre es {cueq:.3f}% CuEq. "
               f"Esta ley CuEq es baja, indicando que el valor económico por tonelada de mineral es limitado. "
               f"La operación dependerá de volúmenes extremadamente altos y costos unitarios bajos. "
               f"Recomendación: Reevaluar la viabilidad económica del proyecto. "
               f"Considerar si los subproductos están siendo recuperados adecuadamente.")
    return txt, level

def interpret_margin(margin, cost_total):
    if margin > 50:
        level = "green"
        txt = (f"El margen operativo es {margin:.2f} USD/t, lo cual representa un {margin/cost_total*100:.1f}% sobre los costos totales ({cost_total:.2f} USD/t). "
               f"Este margen es excelente para una operación minera a cielo abierto. "
               f"Proporciona un amplio colchón ante caídas de precio de los metales o incrementos de costos. "
               f"Recomendación: El proyecto tiene robustez económica. Considerar reinvertir utilidades en exploración para extender la vida de mina.")
    elif margin > 20:
        level = "green"
        txt = (f"El margen operativo es {margin:.2f} USD/t, equivalente al {margin/cost_total*100:.1f}% sobre costos totales ({cost_total:.2f} USD/t). "
               f"Este margen es saludable y típico de operaciones eficientes de mediana a gran escala. "
               f"Permite absorber variaciones moderadas de precio y costo sin comprometer la viabilidad. "
               f"Recomendación: Mantener los controles de costo. Monitorear la tendencia de precios de metales.")
    elif margin > 5:
        level = "yellow"
        txt = (f"El margen operativo es {margin:.2f} USD/t, solo el {margin/cost_total*100:.1f}% sobre costos totales ({cost_total:.2f} USD/t). "
               f"El margen es ajustado. Una caída del precio del cobre o un incremento de costos podría hacer la operación inviable. "
               f"Recomendación: Implementar programas de reducción de costos. Establecer coberturas de precio (hedging) para proteger los ingresos. "
               f"Evaluar la posibilidad de postergar la extracción de bloques marginales.")
    elif margin > 0:
        level = "orange"
        txt = (f"El margen operativo es {margin:.2f} USD/t, apenas el {margin/cost_total*100:.1f}% sobre costos totales ({cost_total:.2f} USD/t). "
               f"La operación es marginalmente rentable. Cualquier incremento de costo o disminución de precio la haría no económica. "
               f"Recomendación: Se requiere acción inmediata para reducir costos o mejorar la recuperación. "
               f"Considerar suspender la extracción de bloques marginales y concentrarse en las zonas de mayor ley.")
    else:
        level = "red"
        txt = (f"El margen operativo es {margin:.2f} USD/t, lo cual es NEGATIVO. "
               f"Los costos totales ({cost_total:.2f} USD/t) superan los ingresos por tonelada. "
               f"La operación NO es económicamente viable bajo las condiciones actuales. "
               f"Recomendación: Suspender la operación hasta que mejoren las condiciones de precio o costo. "
               f"Evaluar alternativas de procesamiento de menor costo.")
    return txt, level

def interpret_rf(rf):
    if rf > 70:
        level = "green"
        txt = (f"El factor de conversión Recurso-Reserva es {rf:.1f}%. "
               f"Esto indica que más del 70% del Recurso Medido+Indicado se convierte en Reserva, lo cual es excelente. "
               f"Refleja un pit de diseño optimizado que captura la mayoría del recurso económico. "
               f"Recomendación: La conversión es muy eficiente. Verificar que los factores modificantes ambientales y sociales no estén subestimados.")
    elif rf > 50:
        level = "green"
        txt = (f"El factor de conversión Recurso-Reserva es {rf:.1f}%. "
               f"Este valor es típico para grandes depósitos a cielo abierto donde limitaciones geotécnicas, ambientales o de pit diseño excluyen parte del recurso. "
               f"Para Antamina, un RF de ~57% sobre Medido+Indicado es consistente con lo reportado. "
               f"Recomendación: El factor es adecuado. Explorar opciones de extensión del pit si los precios mejoran.")
    elif rf > 30:
        level = "yellow"
        txt = (f"El factor de conversión Recurso-Reserva es {rf:.1f}%. "
               f"Este valor indica que una porción significativa del recurso no se convierte en reserva. "
               f"Puede deberse a limitaciones geotécnicas, ribones de pit anchos, o restricciones ambientales. "
               f"Recomendación: Evaluar la optimización del pit para incrementar la conversión. "
               f"Revisar si los ángulos de talud pueden ser más agresivos con estudios geotécnicos adicionales.")
    else:
        level = "orange"
        txt = (f"El factor de conversión Recurso-Reserva es {rf:.1f}%. "
               f"Este valor bajo indica que la mayoría del recurso no se convierte en reserva bajo las condiciones actuales. "
               f"El proyecto puede tener problemas de viabilidad a largo plazo. "
               f"Recomendación: Reevaluar el diseño del pit y los factores modificantes. "
               f"Considerar si cambios en el precio de los metales podrían justificar un pit más profundo.")
    return txt, level

def interpret_dilution(g_dil, g_is, D):
    loss_pct = (g_is - g_dil) / g_is * 100
    if D <= 5:
        level = "green"
        txt = (f"La dilución del {D:.1f}% reduce la ley in situ de {g_is:.3f}% Cu a {g_dil:.3f}% Cu (pérdida del {loss_pct:.2f}%). "
               f"Este nivel de dilución es óptimo para una operación a cielo abierto de gran escala con buen control de voladura y carguío. "
               f"Es consistente con los valores reportados por Antamina (3-8% dilución en tajo abierto). "
               f"Recomendación: Mantener las prácticas actuales de control de dilución.")
    elif D <= 10:
        level = "yellow"
        txt = (f"La dilución del {D:.1f}% reduce la ley in situ de {g_is:.3f}% Cu a {g_dil:.3f}% Cu (pérdida del {loss_pct:.2f}%). "
               f"Este nivel de dilución es moderado. Puede reflejar condiciones geológicas complejas o menor selectividad en el minado. "
               f"Recomendación: Mejorar la definición de los contactos mineral-estéril. "
               f"Optimizar la malla de perforación y el diseño de voladura para reducir la mezcla de material.")
    else:
        level = "orange"
        txt = (f"La dilución del {D:.1f}% reduce la ley in situ de {g_is:.3f}% Cu a {g_dil:.3f}% Cu (pérdida del {loss_pct:.2f}%). "
               f"Este nivel de dilución es alto e impacta significativamente la ley que llega a planta. "
               f"Es típico de operaciones subterráneas o donde el control de contacto es deficiente. "
               f"Recomendación: Implementar programas de control de dilución: GPS en equipos, mejor definición de contactos, "
               f"y monitoreo continuo de la reconciliación mina-planta.")
    return txt, level

def interpret_recoverable(m_rec, m_total, metal_name):
    if m_total == 0:
        return f"No hay contenido metálico de {metal_name} para recuperar.", "green"
    pct = m_rec / m_total * 100
    if pct > 85:
        level = "green"
        txt = (f"Se recuperan {m_rec:,.0f} t de {metal_name} de un total de {m_total:,.0f} t in situ ({pct:.1f}%). "
               f"Esta recuperación es excelente y típica de circuitos de flotación optimizados. "
               f"Recomendación: Mantener las condiciones operativas del circuito de flotación.")
    elif pct > 70:
        level = "green"
        txt = (f"Se recuperan {m_rec:,.0f} t de {metal_name} de un total de {m_total:,.0f} t in situ ({pct:.1f}%). "
               f"Esta recuperación es estándar para operaciones polimetálicas tipo skarn. "
               f"Recomendación: Evaluar la posibilidad de incrementar la recuperación mediante ajustes en el circuito de flotación.")
    elif pct > 50:
        level = "yellow"
        txt = (f"Se recuperan {m_rec:,.0f} t de {metal_name} de un total de {m_total:,.0f} t in situ ({pct:.1f}%). "
               f"La recuperación es moderada. Casi la mitad del metal contenido se pierde en las colas. "
               f"Recomendación: Investigar las causas de la baja recuperación: mineralogía refractaria, granulometría inadecuada, o reactivos subóptimos. "
               f"Considerar ensayos metalúrgicos adicionales.")
    else:
        level = "orange"
        txt = (f"Se recuperan {m_rec:,.0f} t de {metal_name} de un total de {m_total:,.0f} t in situ ({pct:.1f}%). "
               f"La recuperación es baja, indicando que la mayoría del metal contenido no se está recuperando. "
               f"Recomendación: Se requieren estudios metalúrgicos detallados. Evaluar tecnologías alternativas de procesamiento.")
    return txt, level

# ---------------------------- EXCEL EXPORT (in-memory) ----------------------------
def export_excel(p, det, mc, reserve_data):
    wb = openpyxl.Workbook()
    # Sheet 1: Input Data
    ws1 = wb.active
    ws1.title = "1_Datos_Entrada"
    headers1 = ["Parametro / Simbolo", "Valor ingresado", "Unidades", "Valores tipicos por tipo de material / condicion", "CoV Monte Carlo (%)"]
    for c, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.font = XL_HEADER_FONT
        cell.fill = XL_HEADER_FILL
        cell.alignment = XL_ALIGN_C
        cell.border = XL_BORDER
    current_sec = ""
    row = 2
    for k, info in PARAM_INFO.items():
        if info["section"] != current_sec:
            current_sec = info["section"]
            cell = ws1.cell(row=row, column=1, value=current_sec)
            cell.font = XL_SUB_FONT
            cell.fill = XL_SUB_FILL
            cell.border = XL_BORDER
            for cc in range(2, 6):
                ws1.cell(row=row, column=cc).fill = XL_SUB_FILL
                ws1.cell(row=row, column=cc).border = XL_BORDER
            row += 1
        cell = ws1.cell(row=row, column=1, value=info["label"])
        cell.alignment = XL_ALIGN_L
        cell.border = XL_BORDER
        val = p[k]
        cell2 = ws1.cell(row=row, column=2, value=val)
        cell2.border = XL_BORDER
        if isinstance(val, float):
            cell2.number_format = "0.0000"
        cell3 = ws1.cell(row=row, column=3, value=info["unit"])
        cell3.alignment = XL_ALIGN_L
        cell3.border = XL_BORDER
        cell4 = ws1.cell(row=row, column=4, value=info["hint"])
        cell4.alignment = XL_ALIGN_L
        cell4.border = XL_BORDER
        cov = info["std_frac"] * 100 if info["std_frac"] > 0 else 0
        cell5 = ws1.cell(row=row, column=5, value=cov)
        cell5.border = XL_BORDER
        cell5.number_format = "0.00"
        row += 1
    derived = [
        ("Volumen del bloque (V_b)", det["V_b"], "m³", "Calculado automaticamente", ""),
        ("Tonelaje del bloque (T_b)", det["T_b"], "t", "Calculado automaticamente", ""),
        ("Tonelaje total recursos (M+I+Inf)", det["T_total_recursos"], "Mt", "Calculado automaticamente", ""),
        ("Ley diluida (g_dil)", det["g_dil"], "% Cu", "Calculado automaticamente", ""),
        ("Tonelaje ROM (T_ROM)", det["T_ROM"], "t", "Calculado automaticamente", ""),
        ("Tonelaje Reserva reportado (T_R)", det["T_R"], "Mt", "Calculado automaticamente", ""),
        ("Ley de Reserva (g_R)", det["g_R"], "% Cu", "Calculado automaticamente", ""),
        ("Valor total por tonelada (V_total)", det["V_total"], "USD/t", "Calculado automaticamente", ""),
    ]
    cell = ws1.cell(row=row, column=1, value="PARAMETROS DERIVADOS")
    cell.font = XL_SUB_FONT
    cell.fill = XL_SUB_FILL
    cell.border = XL_BORDER
    for cc in range(2, 6):
        ws1.cell(row=row, column=cc).fill = XL_SUB_FILL
        ws1.cell(row=row, column=cc).border = XL_BORDER
    row += 1
    for label, val, unit, tip, cov in derived:
        ws1.cell(row=row, column=1, value=label).alignment = XL_ALIGN_L
        ws1.cell(row=row, column=1).border = XL_BORDER
        cell2 = ws1.cell(row=row, column=2, value=val)
        cell2.border = XL_BORDER
        if isinstance(val, float):
            cell2.number_format = "0.0000"
        ws1.cell(row=row, column=3, value=unit).border = XL_BORDER
        ws1.cell(row=row, column=4, value=tip).alignment = XL_ALIGN_L
        ws1.cell(row=row, column=4).border = XL_BORDER
        ws1.cell(row=row, column=5, value=cov).border = XL_BORDER
        row += 1
    ws1.column_dimensions["A"].width = 42
    ws1.column_dimensions["B"].width = 18
    ws1.column_dimensions["C"].width = 12
    ws1.column_dimensions["D"].width = 55
    ws1.column_dimensions["E"].width = 22

    # Sheet 2: Interpretation
    ws2 = wb.create_sheet("2_Interpretacion_Resultados")
    headers2 = ["Resultado", "Valor", "Unidades", "Interpretacion Ingenieril"]
    for c, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = XL_HEADER_FONT
        cell.fill = XL_HEADER_FILL
        cell.alignment = XL_ALIGN_C
        cell.border = XL_BORDER
    interp_items = [
        ("Ley de Corte (COG)", f"{det['COG']:.4f}", "% Cu", *interpret_cog(det["COG"])),
        ("Equivalente de Cobre (CuEq)", f"{det['CuEq']:.3f}", "% CuEq", *interpret_cueq(det["CuEq"])),
        ("Margen Operativo", f"{det['margin']:.2f}", "USD/t", *interpret_margin(det["margin"], det["cost_total"])),
        ("Factor Conversion Recurso-Reserva", f"{det['RF']:.1f}", "%", *interpret_rf(det["RF"])),
        ("Ley Diluida", f"{det['g_dil']:.4f}", "% Cu", *interpret_dilution(det["g_dil"], p["g_Cu"], p["D"])),
    ]
    m_cu_total = calc_metal_content_base(p["T_res_Cu_ore_P"] * 1e6, p["g_Cu"])
    m_cu_rec = det["M_Cu_rec"]
    interp_items.append(("Cobre Recuperable (Prob. Cu-ore)", f"{m_cu_rec:,.0f}", "t Cu", *interpret_recoverable(m_cu_rec, m_cu_total, "Cu")))
    row2 = 2
    for label, val, unit, txt, level in interp_items:
        ws2.cell(row=row2, column=1, value=label).alignment = XL_ALIGN_L
        ws2.cell(row=row2, column=1).border = XL_BORDER
        ws2.cell(row=row2, column=2, value=val).border = XL_BORDER
        ws2.cell(row=row2, column=3, value=unit).border = XL_BORDER
        cell4 = ws2.cell(row=row2, column=4, value=txt)
        cell4.alignment = XL_ALIGN_L
        cell4.border = XL_BORDER
        if level == "green":
            cell4.fill = XL_GREEN
        elif level == "yellow":
            cell4.fill = XL_YELLOW
        elif level == "orange":
            cell4.fill = XL_ORANGE
        elif level == "red":
            cell4.fill = XL_RED
        row2 += 1
    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 90

    # Sheet 3: Step-by-step
    ws3 = wb.create_sheet("3_Calculo_JORC_NI43101")
    headers3 = ["Modulo", "Parametro Calculado", "Formula", "Valores Utilizados", "Resultado", "Unidad"]
    for c, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = XL_HEADER_FONT
        cell.fill = XL_HEADER_FILL
        cell.alignment = XL_ALIGN_C
        cell.border = XL_BORDER
    steps = [
        ("Mod 2", "Volumen del bloque (V_b)", "V_b = x * y * z", f"{p['x']}*{p['y']}*{p['z']}", det["V_b"], "m³"),
        ("Mod 2", "Tonelaje del bloque (T_b)", "T_b = V_b * rho_b", f"{det['V_b']}*{p['rho_b']}", det["T_b"], "t"),
        ("Mod 3", "Contenido Cu in situ (M_Cu)", "M_Cu = T_total * g_Cu / 100", f"{p['T_res_Cu_ore_P']*1e6:.0f} * {p['g_Cu']} / 100", det["M_Cu_total"], "t Cu"),
        ("Mod 3", "Contenido Ag in situ (M_Ag)", "M_Ag = T_total * g_Ag / 1e6", f"{p['T_res_Cu_ore_P']*1e6:.0f} * {p['g_Ag']} / 1e6", det["M_Ag_total"], "t Ag"),
        ("Mod 4", "Costo total (C_total)", "C_total = C_min + C_proc + C_vta", f"{p['C_min']}+{p['C_proc']}+{p['C_vta']}", det["cost_total"], "USD/t"),
        ("Mod 4", "Ley de Corte (COG)", "COG = C_total / (P_Cu*R_Cu*(1-r))", f"{det['cost_total']:.2f} / ({p['P_Cu']}*{p['R_Cu']}*{1-p['r_royalty']:.2f})", det["COG"], "% Cu"),
        ("Mod 5", "Ley diluida (g_dil)", "g_dil = g_is*(1-D/100) + g_w*(D/100)", f"{p['g_Cu']}*(1-{p['D']/100}) + {p['g_w']}*({p['D']/100})", det["g_dil"], "% Cu"),
        ("Mod 5", "Tonelaje ROM (T_ROM)", "T_ROM = T_res * MR / (1-D/100)", f"{p['T_res_Cu_ore_P']*1e6:.0f} * {p['MR']} / (1-{p['D']/100})", det["T_ROM"], "t"),
        ("Mod 6", "Dilucion absoluta (D_abs)", "D_abs = D / (100-D)", f"{p['D']} / (100-{p['D']})", det["D_abs"], "fraccion"),
        ("Mod 6", "Tonelaje Reserva (T_R)", "T_R = T_res * MR * (1+D_abs)", f"{p['T_res_Cu_ore_P']*1e6:.0f} * {p['MR']} * (1+{det['D_abs']:.6f})", det["T_R"], "t"),
        ("Mod 6", "Ley de Reserva (g_R)", "g_R = g_dil (simplificado)", "Igual a g_dil", det["g_R"], "% Cu"),
        ("Mod 7", "Cu recuperable (M_Cu_rec)", "M_Cu_rec = T_R * g_R / 100 * R_Cu", f"{det['T_R']:.0f} * {det['g_R']:.4f} / 100 * {p['R_Cu']}", det["M_Cu_rec"], "t Cu"),
        ("Mod 7", "Ag recuperable (M_Ag_rec)", "M_Ag_rec = T_R * g_Ag / 1e6 * R_Ag", f"{det['T_R']:.0f} * {p['g_Ag']} / 1e6 * {p['R_Ag']}", det["M_Ag_rec"], "t Ag"),
        ("Mod 8", "Valor Cu/t (V_Cu)", "V_Cu = g_Cu*0.01*P_Cu*R_Cu", f"{p['g_Cu']}*0.01*{p['P_Cu']}*{p['R_Cu']}", det["V_Cu"], "USD/t"),
        ("Mod 8", "Valor Zn/t (V_Zn)", "V_Zn = g_Zn*0.01*P_Zn*R_Zn", f"{p['g_Zn']}*0.01*{p['P_Zn']}*{p['R_Zn']}", det["V_Zn"], "USD/t"),
        ("Mod 8", "Valor Ag/t (V_Ag)", "V_Ag = g_Ag*1e-6*P_Ag*R_Ag", f"{p['g_Ag']}*1e-6*{p['P_Ag']}*{p['R_Ag']}", det["V_Ag"], "USD/t"),
        ("Mod 8", "Valor Mo/t (V_Mo)", "V_Mo = g_Mo*0.01*P_Mo*R_Mo", f"{p['g_Mo']}*0.01*{p['P_Mo']}*{p['R_Mo']}", det["V_Mo"], "USD/t"),
        ("Mod 8", "Valor total/t (V_total)", "V_total = V_Cu+V_Zn+V_Ag+V_Mo", "Suma de valores individuales", det["V_total"], "USD/t"),
        ("Mod 8", "Equivalente Cobre (CuEq)", "CuEq = V_total / (P_Cu*R_Cu*0.01)", f"{det['V_total']:.2f} / ({p['P_Cu']}*{p['R_Cu']}*0.01)", det["CuEq"], "% CuEq"),
        ("Mod 9", "Factor Conversion (RF)", "RF = T_Reserva / T_Recurso_MI * 100", f"{p['T_res_Cu_ore_P']+p['T_res_CuZn_ore_P']+p['T_res_Cu_ore_Pr']+p['T_res_CuZn_ore_Pr']} / {p['T_Recurso_Medido']+p['T_Recurso_Indicado']} * 100", det["RF"], "%"),
        ("Mod 9", "Ingreso por tonelada (I_t)", "I_t = CuEq/100 * P_Cu * R_Cu", f"{det['CuEq']}/100 * {p['P_Cu']} * {p['R_Cu']}", det["I_t"], "USD/t"),
        ("Mod 9", "Margen operativo", "Margen = I_t - C_total", f"{det['I_t']:.2f} - {det['cost_total']:.2f}", det["margin"], "USD/t"),
    ]
    for i, (mod, param, formula, vals, result, unit) in enumerate(steps, 2):
        ws3.cell(row=i, column=1, value=mod).alignment = XL_ALIGN_C
        ws3.cell(row=i, column=1).border = XL_BORDER
        ws3.cell(row=i, column=2, value=param).border = XL_BORDER
        ws3.cell(row=i, column=3, value=formula).border = XL_BORDER
        ws3.cell(row=i, column=4, value=vals).border = XL_BORDER
        cell_res = ws3.cell(row=i, column=5, value=result)
        cell_res.border = XL_BORDER
        if isinstance(result, float):
            cell_res.number_format = "0.0000"
        ws3.cell(row=i, column=6, value=unit).border = XL_BORDER
    ws3.column_dimensions["A"].width = 10
    ws3.column_dimensions["B"].width = 32
    ws3.column_dimensions["C"].width = 42
    ws3.column_dimensions["D"].width = 45
    ws3.column_dimensions["E"].width = 18
    ws3.column_dimensions["F"].width = 12

    # Sheet 4: Monte Carlo Statistics (simplified)
    ws4 = wb.create_sheet("4_MonteCarlo_Estadisticas")
    # We'll just add some key stats (since full MC sample may be huge)
    ws4.cell(row=1, column=1, value="ESTADISTICAS DE VARIABLES DE SALIDA").font = XL_SUB_FONT
    ws4.cell(row=1, column=1).fill = XL_SUB_FILL
    for c in [2,3,4,5,6,7]:
        ws4.cell(row=1, column=c).fill = XL_SUB_FILL
    headers4 = ["Variable", "Media", "Desv. Est.", "P5", "P50", "P90", "P95"]
    for c, h in enumerate(headers4, 1):
        cell = ws4.cell(row=2, column=c, value=h)
        cell.font = XL_HEADER_FONT
        cell.fill = XL_HEADER_FILL
        cell.alignment = XL_ALIGN_C
        cell.border = XL_BORDER
    mc_var_names = {"COG": "% Cu", "CuEq": "% CuEq", "margin": "USD/t", "M_Cu_rec": "t Cu"}
    row4 = 3
    for name, unit in mc_var_names.items():
        sts = mc_stats(mc[name])
        ws4.cell(row=row4, column=1, value=f"{name} ({unit})").border = XL_BORDER
        ws4.cell(row=row4, column=2, value=sts["mean"]).border = XL_BORDER
        ws4.cell(row=row4, column=3, value=sts["std"]).border = XL_BORDER
        ws4.cell(row=row4, column=4, value=sts["P5"]).border = XL_BORDER
        ws4.cell(row=row4, column=5, value=sts["P50"]).border = XL_BORDER
        ws4.cell(row=row4, column=6, value=sts["P90"]).border = XL_BORDER
        ws4.cell(row=row4, column=7, value=sts["P95"]).border = XL_BORDER
        row4 += 1
    ws4.column_dimensions["A"].width = 25
    for i in range(2,8):
        ws4.column_dimensions[get_column_letter(i)].width = 16

    wb_data = io.BytesIO()
    wb.save(wb_data)
    wb_data.seek(0)
    return wb_data

# ---------------------------- MAIN STREAMLIT APP ----------------------------
def main():
    st.set_page_config(page_title="JORC/NI 43-101 Economic Analysis", layout="wide")
    st.title("📊 JORC / NI 43-101 Economic Analysis")
    st.markdown("Interactive tool for mineral resource and reserve economic evaluation (Cu-Zn-Ag-Mo skarn / porphyry).")

    # Sidebar inputs
    with st.sidebar:
        st.header("Input Parameters")
        with st.expander("ℹ️ Instructions", expanded=False):
            st.markdown("Enter parameters for each section. Default values are based on typical skarn deposits. Use the 'Run Analysis' button at the bottom to compute results.")
        # Group inputs by section
        sections = {}
        for k, info in PARAM_INFO.items():
            sec = info["section"]
            if sec not in sections:
                sections[sec] = []
            sections[sec].append((k, info))
        # Create a form to group all inputs and avoid constant recomputation
        with st.form("inputs_form"):
            user_params = {}
            for sec, items in sections.items():
                st.subheader(sec)
                col1, col2 = st.columns(2)
                for i, (k, info) in enumerate(items):
                    if k == "N_sim":
                        # simulation number input as integer
                        val = st.number_input(
                            f"{info['label']}",
                            value=int(info["default"]),
                            step=1000,
                            help=info["hint"],
                            key=k
                        )
                    else:
                        val = st.number_input(
                            f"{info['label']} ( {info['unit']} )",
                            value=float(info["default"]),
                            step=0.1 if "price" not in k.lower() else 100.0,
                            format="%.4f" if info["unit"] != "-" else "%.0f",
                            help=info["hint"],
                            key=k
                        )
                    user_params[k] = val
                st.markdown("---")
            submitted = st.form_submit_button("🚀 Run Analysis")
            if not submitted:
                st.stop()
        # After submit, proceed

    # Run analysis only when button is pressed
    with st.spinner("Running deterministic and Monte Carlo calculations..."):
        # Deterministic calculations (same as original run but without plots)
        p = user_params
        N_sim = int(p["N_sim"])
        V_b, T_b = calc_block_tonnage(p["x"], p["y"], p["z"], p["rho_b"])
        T_total_recursos = p["T_Recurso_Medido"] + p["T_Recurso_Indicado"] + p["T_Recurso_Inferido"]
        T_ref = p["T_res_Cu_ore_P"] * 1e6
        M_Cu_total = calc_metal_content_base(T_ref, p["g_Cu"])
        M_Zn_total = calc_metal_content_base(T_ref, p["g_Zn"])
        M_Ag_total = calc_metal_content_precious(T_ref, p["g_Ag"])
        M_Mo_total = calc_metal_content_base(T_ref, p["g_Mo"])
        COG, cost_total = calc_cog(p["C_min"], p["C_proc"], p["C_vta"], p["P_Cu"], p["R_Cu"], p["r_royalty"])
        g_dil = calc_diluted_grade(p["g_Cu"], p["D"], p["g_w"])
        T_ROM = calc_rom_tonnage(T_ref, p["D"], p["MR"])
        D_abs = p["D"] / (100 - p["D"]) if (100 - p["D"]) > 0 else 0
        T_R = calc_reserve_tonnage(T_ref, p["MR"], p["D"])
        g_R = calc_diluted_grade(p["g_Cu"], p["D"], p["g_w"])
        M_Cu_rec = calc_recoverable_metal_base(T_R, g_R, p["R_Cu"])
        M_Ag_rec = calc_recoverable_metal_precious(T_R, p["g_Ag"], p["R_Ag"])
        CuEq, V_total, V_Cu, V_Zn, V_Ag, V_Mo = calc_cueq(
            p["g_Cu"], p["g_Zn"], p["g_Ag"], p["g_Mo"],
            p["P_Cu"], p["P_Zn"], p["P_Ag"], p["P_Mo"],
            p["R_Cu"], p["R_Zn"], p["R_Ag"], p["R_Mo"],
        )
        T_Recurso_MI = p["T_Recurso_Medido"] + p["T_Recurso_Indicado"]
        T_Reserva_total = p["T_res_Cu_ore_P"] + p["T_res_CuZn_ore_P"] + p["T_res_Cu_ore_Pr"] + p["T_res_CuZn_ore_Pr"]
        RF = calc_rf(T_Reserva_total, T_Recurso_MI)
        I_t, margin, _ = calc_margin(CuEq, p["P_Cu"], p["R_Cu"], p["C_min"], p["C_proc"], p["C_vta"])
        det = {
            "V_b": V_b, "T_b": T_b, "T_total_recursos": T_total_recursos,
            "M_Cu_total": M_Cu_total, "M_Zn_total": M_Zn_total,
            "M_Ag_total": M_Ag_total, "M_Mo_total": M_Mo_total,
            "COG": COG, "cost_total": cost_total,
            "g_dil": g_dil, "T_ROM": T_ROM,
            "D_abs": D_abs, "T_R": T_R / 1e6, "g_R": g_R,
            "M_Cu_rec": M_Cu_rec, "M_Ag_rec": M_Ag_rec,
            "CuEq": CuEq, "V_total": V_total, "V_Cu": V_Cu, "V_Zn": V_Zn, "V_Ag": V_Ag, "V_Mo": V_Mo,
            "RF": RF, "I_t": I_t, "margin": margin,
        }

        # Reserve data for plots
        reserve_data = []
        for name, T_key, gcu, gzn, gag, gmo in [
            ("Probada Cu-ore", "T_res_Cu_ore_P", p["g_Cu"], p["g_Zn"], p["g_Ag"], p["g_Mo"]),
            ("Probada Cu-Zn", "T_res_CuZn_ore_P", p["g_Cu_CuZn_P"], p["g_Zn_CuZn_P"], p["g_Ag_CuZn_P"], p["g_Mo_CuZn_P"]),
            ("Probable Cu-ore", "T_res_Cu_ore_Pr", p["g_Cu_Cu_Pr"], p["g_Zn_Cu_Pr"], p["g_Ag_Cu_Pr"], p["g_Mo_Cu_Pr"]),
            ("Probable Cu-Zn", "T_res_CuZn_ore_Pr", p["g_Cu_CuZn_Pr"], p["g_Zn_CuZn_Pr"], p["g_Ag_CuZn_Pr"], p["g_Mo_CuZn_Pr"]),
        ]:
            T_val = p[T_key] * 1e6
            m_cu = calc_metal_content_base(T_val, gcu)
            m_zn = calc_metal_content_base(T_val, gzn)
            m_ag = calc_metal_content_precious(T_val, gag)
            m_mo = calc_metal_content_base(T_val, gmo)
            g_dil_cat = calc_diluted_grade(gcu, p["D"], p["g_w"])
            T_R_cat = calc_reserve_tonnage(T_val, p["MR"], p["D"])
            m_cu_rec = calc_recoverable_metal_base(T_R_cat, g_dil_cat, p["R_Cu"])
            m_zn_rec = calc_recoverable_metal_base(T_R_cat, g_dil_cat if gzn < 1 else calc_diluted_grade(gzn, p["D"], 0), p["R_Zn"])
            m_ag_rec = calc_recoverable_metal_precious(T_R_cat, gag, p["R_Ag"])
            m_mo_rec = calc_recoverable_metal_base(T_R_cat, calc_diluted_grade(gmo, p["D"], 0), p["R_Mo"])
            reserve_data.append({
                "name": name,
                "T": p[T_key],
                "g_Cu": gcu, "g_Zn": gzn, "g_Ag": gag, "g_Mo": gmo,
                "M_Cu": m_cu, "M_Zn": m_zn, "M_Ag": m_ag, "M_Mo": m_mo,
                "M_Cu_rec": m_cu_rec, "M_Zn_rec": m_zn_rec, "M_Ag_rec": m_ag_rec, "M_Mo_rec": m_mo_rec,
            })

        # Monte Carlo
        mc = monte_carlo(p, N_sim)

    # ---- Display Results ----
    st.header("Deterministic Results")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Cut‑Off Grade (COG)", f"{det['COG']:.4f}% Cu")
        st.metric("Copper Equivalent (CuEq)", f"{det['CuEq']:.3f}%")
        st.metric("Operating Margin", f"{det['margin']:.2f} USD/t")
    with col2:
        st.metric("Diluted Grade", f"{det['g_dil']:.4f}% Cu")
        st.metric("Reserve Tonnage", f"{det['T_R']:.2f} Mt")
        st.metric("Resource → Reserve Factor", f"{det['RF']:.1f}%")
    with col3:
        st.metric("Total Resources (M+I+Inf)", f"{det['T_total_recursos']:.0f} Mt")
        st.metric("Recoverable Copper", f"{det['M_Cu_rec']:,.0f} t")
        st.metric("Revenue per tonne", f"{det['I_t']:.2f} USD/t")

    st.subheader("Interpretive Comments")
    # COG interpretation
    txt_cog, level_cog = interpret_cog(det["COG"])
    st.info(f"{risk_fill(level_cog)} **COG**: {txt_cog}")
    # CuEq interpretation
    txt_cueq, level_cueq = interpret_cueq(det["CuEq"])
    st.info(f"{risk_fill(level_cueq)} **CuEq**: {txt_cueq}")
    # Margin
    txt_margin, level_margin = interpret_margin(det["margin"], det["cost_total"])
    st.info(f"{risk_fill(level_margin)} **Margin**: {txt_margin}")
    # RF
    txt_rf, level_rf = interpret_rf(det["RF"])
    st.info(f"{risk_fill(level_rf)} **Resource‑Reserve Conversion**: {txt_rf}")
    # Dilution
    txt_dil, level_dil = interpret_dilution(det["g_dil"], p["g_Cu"], p["D"])
    st.info(f"{risk_fill(level_dil)} **Dilution**: {txt_dil}")

    st.header("Monte Carlo Results")
    st.write(f"Based on {N_sim:,} simulations.")
    # Statistics table
    mc_vars = ["COG", "CuEq", "margin", "M_Cu_rec"]
    stats_df = pd.DataFrame({
        "Variable": ["COG (% Cu)", "CuEq (% CuEq)", "Margin (USD/t)", "Recoverable Cu (t)"],
        "Mean": [mc_stats(mc[v])["mean"] for v in mc_vars],
        "Std Dev": [mc_stats(mc[v])["std"] for v in mc_vars],
        "P5": [mc_stats(mc[v])["P5"] for v in mc_vars],
        "P50": [mc_stats(mc[v])["P50"] for v in mc_vars],
        "P95": [mc_stats(mc[v])["P95"] for v in mc_vars],
    })
    st.dataframe(stats_df.style.format("{:.4f}"))

    # Histograms
    st.subheader("Probability Distributions")
    col_hist1, col_hist2 = st.columns(2)
    with col_hist1:
        fig1 = plot_hist_cdf(mc["COG"], "Cut-Off Grade (COG)", "COG (% Cu)")
        st.pyplot(fig1)
        plt.close(fig1)
        fig2 = plot_hist_cdf(mc["margin"], "Operating Margin", "Margin (USD/t)")
        st.pyplot(fig2)
        plt.close(fig2)
    with col_hist2:
        fig3 = plot_hist_cdf(mc["CuEq"], "Copper Equivalent", "CuEq (%)")
        st.pyplot(fig3)
        plt.close(fig3)
        fig4 = plot_hist_cdf(mc["M_Cu_rec"] / 1e6, "Recoverable Copper", "M_Cu_rec (Mt Cu)")
        st.pyplot(fig4)
        plt.close(fig4)

    st.subheader("Sensitivity Analysis")
    # Tornado
    fig_tornado = plot_tornado(mc, p)
    st.pyplot(fig_tornado)
    plt.close(fig_tornado)

    # Scatter plots
    col_scat1, col_scat2 = st.columns(2)
    with col_scat1:
        fig_scat1 = plot_scatter(mc, "P_Cu", "margin", "Precio Cu (USD/t)", "Margen Operativo (USD/t)", "Precio Cu vs Margen")
        st.pyplot(fig_scat1)
        plt.close(fig_scat1)
        fig_scat2 = plot_scatter(mc, "g_Cu", "CuEq", "Ley Cu (%)", "CuEq (%)", "Ley Cu vs CuEq")
        st.pyplot(fig_scat2)
        plt.close(fig_scat2)
    with col_scat2:
        fig_scat3 = plot_scatter(mc, "C_proc", "COG", "Costo Procesamiento (USD/t)", "COG (% Cu)", "Costo Procesamiento vs COG")
        st.pyplot(fig_scat3)
        plt.close(fig_scat3)

    # Deterministic sensitivity plots
    st.subheader("Deterministic Sensitivity")
    col_sens1, col_sens2 = st.columns(2)
    with col_sens1:
        fig_sens_price = plot_det_sensitivity_price(p)
        st.pyplot(fig_sens_price)
        plt.close(fig_sens_price)
    with col_sens2:
        fig_sens_grade = plot_det_sensitivity_grade(p)
        st.pyplot(fig_sens_grade)
        plt.close(fig_sens_grade)

    # Resource/Reserve plots
    st.subheader("Resource & Reserve Graphics")
    col_res1, col_res2 = st.columns(2)
    with col_res1:
        fig_pie = plot_resource_pie(p)
        st.pyplot(fig_pie)
        plt.close(fig_pie)
        fig_comp = plot_resource_reserve_comparison(p)
        st.pyplot(fig_comp)
        plt.close(fig_comp)
    with col_res2:
        fig_bar = plot_metal_content_bar(reserve_data)
        st.pyplot(fig_bar)
        plt.close(fig_bar)
        fig_cueq_cat = plot_cueq_by_category(p)
        st.pyplot(fig_cueq_cat)
        plt.close(fig_cueq_cat)

    # Excel download
    st.subheader("Download Report")
    excel_data = export_excel(p, det, mc, reserve_data)
    st.download_button(
        label="📥 Download Excel Report (JORC/NI 43-101)",
        data=excel_data,
        file_name="jorc_ni43-101_report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

if __name__ == "__main__":
    main()